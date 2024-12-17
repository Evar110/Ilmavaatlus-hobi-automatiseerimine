################################################
# Programmeerimine I
# 2024/2025 sügissemester
#
# Projekt
# Teema:
# Ilmavaatlus hobi automatiseerimine
#
# Autorid:
# Kristjian Caius Kasuk
# Evar Valentin Pereseld
#
# Lisakommentaar:
# Vaja on installida Pythoni moodulid tkinter, tkcalendar, bs4, selenium, openpyxl.
# Oleks vaja ka tõmmata alla geckodriver(https://github.com/mozilla/geckodriver/releases) ja lisada koodi reale 344, selle .exe asukoht arvutis.
# Ilma andmete allikas on Keskkonnaagentuur(https://www.ilmateenistus.ee/ilm/ilmavaatlused/vaatlusandmed/tunniandmed/)
##################################################

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.firefox.service import Service as FirefoxService
from bs4 import BeautifulSoup
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Side, Font
from tkinter import Tk, Label, Entry, Button, Toplevel, filedialog, ttk, messagebox, PhotoImage
from tkcalendar import Calendar
import shutil
import os

# funktsioon andmete kogumiseks
def andmete_kogumine(kellaaeg, kogutav, koht):
    """
    :parameeter kellaaeg: Kellaaeg, mida kasutada päringus (näiteks "7:00").
    :parameeter kogutav: Andmed, mida koguda ('hommik' või 'õhtu').
    :parameeter koht: Asukoht, kus andmeid võtab.
    """
    global driver
    try:
        # Avab veebilehe
        driver.get(url)

        # Kuupäeva sisestamine
        date_input = WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.ID, "ood-datepicker"))) # ootab 20s või kuni saab lisada kuupäeva
        date_input.clear()
        date_input.send_keys(kuupäev)
        date_input.send_keys(Keys.RETURN)
        
        # Kellaaja sisestamine
        time_input = WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.ID, "ood-timepicker"))) # ootab 20s või kuni saab lisada kellaaja
        time_input.clear()
        time_input.send_keys(kellaaeg)
        time_input.send_keys(Keys.RETURN)

        # Ootab 20s või kuni leht lõpetab andmete laadimise
        WebDriverWait(driver, 20).until(lambda d: driver.execute_script("return jQuery.active == 0"))

        # Parsib HTML-i BeautifulSoup abil
        html = driver.page_source
        soup = BeautifulSoup(html, "html.parser")
        tabel = soup.find("table", class_="table")

        if tabel:
            read = tabel.find_all("tr")[1:]  # Jätab vahele päise rea
            for rida in read:
                andmete_rida = rida.find_all("td")
                if andmete_rida:
                    asukoht = andmete_rida[0].get_text(strip=True)
                    temperatuur = andmete_rida[1].get_text(strip=True)
                    õhurõhk = andmete_rida[3].get_text(strip=True)
                    tuule_suund = andmete_rida[5].get_text(strip=True)

                    if asukoht == koht:  # Filtreerib valitud asukoha andmed
                        if kogutav == "hommik":
                            print(f"Asukoht: {asukoht}")
                        print(f"Temperatuur: {temperatuur if temperatuur else '-'}")
                        andmed.append(temperatuur if temperatuur else '-')

                        if kogutav == "õhtu":
                            print(f"Õhurõhk: {õhurõhk if õhurõhk else '-'}")
                            print(f"Tuule suund: {tuule_suund if tuule_suund else '-'}")
                            andmed.extend([
                                õhurõhk if õhurõhk else '-',
                                tuule_suund if tuule_suund else '-'
                            ])
                            print("-" * 40)
    except Exception as e:
        print(f"Tekkis viga: {e}")

# funktsioon nädalapäeva saamiseks
def nädalapäev(kuupäev):
    nädalapäevad = ['E', 'T', 'K', 'N', 'R', 'L', 'P']
    
    kuupäeva_obj = datetime.strptime(kuupäev, "%d.%m.%Y")
    nädalapäeva_indeks = kuupäeva_obj.weekday()
    return f"{kuupäeva_obj.day}.{nädalapäevad[nädalapäeva_indeks]}."

# funktsioon, mis teisendab hektopaskalid mmHg-ks
def hPa_mmHg(hPa):
    if hPa == '-':
        return hPa
    else:
        mmHg = str(round(float(hPa) * 0.750062)) + "mmHg"
        return mmHg
    

# funktsioon, mis muudab tuule suuna kraad ilmakaareks
def tuule_suund(kraadid):
    ilmakaared = ["Põhja", "Kirde", "Ida", "Kagu", "Lõuna", "Edela", "Lääne", "Loode", "Põhja"]
    if kraadid == '-':
            return kraadid
    else:    
        suuna_indeks = int((int(kraadid) + 22.5) % 360 // 45)
        return ilmakaared[suuna_indeks]

# Funktsioon, mis kontrollib kas on uus aasta
def uus_aasta(kuupäev):
    if int(kuupäev[:2]) == 1 and int(kuupäev[3:5]) == 1:
        return True
    else:
        return False
    
# funktsioon, mis kontrollib kas on uus kuu
def uus_kuu(kuupäev):
    if int(kuupäev[:2]) == 1:
        return True
    else:
        return False

# funktsioon, mis tagastab kuu ja aasta, mida kasutaja sisestas    
def kuu_ja_aasta(kuupäev):
    järjend = []
    kuud = ['Jaanuar', 'Veebruar', 'Märts', 'Aprill', 'Mai', 'Juuni', 'Juuli', 'August', 'September', 'Oktoober', 'November', 'Detsember']
    kuu = kuud[int(kuupäev[3:5]) - 1]
    järjend.append(kuu)
    aasta = int(kuupäev[6:])
    järjend.append(aasta)
    return järjend
    
# funktsioon, mis kirjutab andmed Exceli faili
def lisa_excelisse(failinimi, andmed, kuupäev):
    """
    Lisab andmed Exceli faili uuele reale, kui faili ei ole, loob selle.
    Lisaks määrab joondamise, stiili ja muud visuaalsed elemendid.
    """
    try:
        # Kuu ja aasta
        kuu_ja_aasta_järjend = kuu_ja_aasta(kuupäev)
        aasta = kuu_ja_aasta_järjend[1]
        kuu = kuu_ja_aasta_järjend[0]

        # Stiilid
        keskele_joondus = Alignment(horizontal="center")
        alläär = Border(bottom=Side(style="thin"))
        suur_font = Font(size=18)

        # Spetsiaalsed read
        päis_rida = ["Kuupäev", "Hommiku t°", "Õhtu t°", "Õhurõhk", "Tuule suund"]
        kuu_rida = ["", "", kuu]
        aasta_rida = ["", "", aasta]

        #Juhul, kui ei ole sellist faili, loob selle
        if not os.path.exists(failinimi):
            fail = Workbook()
            excel = fail.active

            # Lisab aasta
            excel.append(aasta_rida)
            for lahter in excel[excel.max_row]:
                lahter.alignment = keskele_joondus
                lahter.font = suur_font

            # Lisab kuu
            excel.append(kuu_rida)
            for lahter in excel[excel.max_row]:
                lahter.alignment = keskele_joondus

            # Lisab päise
            excel.append(päis_rida)
            for lahter in excel[excel.max_row]:
                lahter.alignment = keskele_joondus
                lahter.border = alläär

            # Lisab andmed
            excel.append(andmed)
            for lahter in excel[excel.max_row]:
                lahter.alignment = keskele_joondus

            # Salvestab andmed
            fail.save(failinimi)

        else:
            fail = load_workbook(failinimi)
            excel = fail.active
            # Kui on uus aasta, lisab uue aasta, kuu ja päise
            if uus_aasta(kuupäev):
                excel.append(aasta_rida)
                for lahter in excel[excel.max_row]:
                    lahter.alignment = keskele_joondus
                    lahter.font = suur_font

                excel.append(kuu_rida)
                for lahter in excel[excel.max_row]:
                    lahter.alignment = keskele_joondus

                excel.append(päis_rida)
                for lahter in excel[excel.max_row]:
                    lahter.alignment = keskele_joondus
                    lahter.border = alläär

            # Kui on uus kuu, lisab uue kuu
            elif uus_kuu(kuupäev):
                excel.append(kuu_rida)
                for lahter in excel[excel.max_row]:
                    lahter.alignment = keskele_joondus

            # Lisab andmed        
            excel.append(andmed)
            for lahter in excel[excel.max_row]:
                lahter.alignment = keskele_joondus
        
            # Salvestab andmed
            fail.save(failinimi)

    # Juhul, kui kasutaja jätab exceli faili lahti, kui programmi käivitab
    except PermissionError:
        print("Viga, ei saanud salvestada faili, kontrolli ega sul exceli fail avatud ei ole")

# Funktsioon, mis kustutab varasemad kogutud andmed, et programm saaks uuesti koguda andmeid, ilma restarti tegematta.
def restart():
    global andmed
    andmed = []

def käivita():
    global kuupäev, hommikune_aeg, õhtune_aeg, asukoht, andmed, driver

    kuupäev = kuupäev_sisend.get()
    hommikune_aeg = hommikune_aeg_sisend.get()
    õhtune_aeg = õhtune_aeg_sisend.get()
    asukoht = asukoht_sisend.get()

    # Kontroll
    if len(kuupäev) != 10 or not hommikune_aeg or not õhtune_aeg or not asukoht:
        print("Palun täida kõik väljad!")
        return
    
    restart()
    
    # Käivitab Firefox WebDriver
    service = FirefoxService(executable_path=geckodriver_path)
    driver = webdriver.Firefox(service=service)

    try:
        # Andmete kogumine hommikul (ainult temperatuur)
        andmete_kogumine(hommikune_aeg, "hommik", asukoht)

        # Andmete kogumine õhtul (temperatuur, õhurõhk, tuule suund)
        andmete_kogumine(õhtune_aeg, "õhtu", asukoht)
    finally:
        driver.quit()

    # Muudab andmed järjendi elemente, et vastaksid tahetud formaadile
    päev_nädalapäev = nädalapäev(kuupäev)
    andmed.insert(0, päev_nädalapäev)
    õhurõhk = andmed[3]
    õhurõhk = õhurõhk.replace(',','.')
    õhurõhk = hPa_mmHg(õhurõhk)
    andmed[3] = õhurõhk
    tuule_kraadid = andmed[4]
    ilmakaar = tuule_suund(tuule_kraadid)
    andmed[4] = ilmakaar

    print("Kogutud andmed:", andmed)

    failinimi = "ilm.xlsx"
    lisa_excelisse(failinimi, andmed, kuupäev)

    print(f"Andmed salvestatud faili {failinimi}.")

# Funktsioon kuupäeva valimiseks
def vali_kuupäev():
    def valitud_kuupäev():
        kuupäev_sisend.delete(0, 'end')
        kuupäev_sisend.insert(0, cal.get_date())
        calendar_window.destroy()

    calendar_window = Toplevel(root)
    calendar_window.title("Vali kuupäev")
    cal = Calendar(calendar_window, selectmode='day', date_pattern='dd.mm.yyyy')
    cal.pack(pady=20)

    Button(calendar_window, text="Vali", command=valitud_kuupäev).pack(pady=10)

# Asukohad
asukohad = [
    "Ahja", "Audru", "Dirhami", "Haapsalu", "Haapsalu sadam", "Heltermaa", "Häädemeeste",
    "Hüüru", "Jõgeva", "Jõhvi", "Kaansoo", "Kasari", "Kassari", "Kehra", "Keila", "Kihnu",
    "Kloostrimetsa", "Koodu", "Korela", "Kulgu sadam", "Kunda", "Kuningaküla", "Kuressaare linn",
    "Kuusiku", "Kõrgessaare", "Laadi", "Laimjala", "Loksa", "Luguse", "Lüganuse", "Lääne-Nigula",
    "Mehikoorma", "Mustvee", "Mõntu", "Männikjärve raba", "Naissaare", "Narva", "Narva Karjääri",
    "Narva linn", "Nurme", "Oore", "Osmussaare", "Otepää", "Pajupea", "Pajusi", "Paldiski (Põhjasadam)",
    "Pakri", "Piigaste", "Pirita", "Praaga", "Pärnu", "Rannu-Jõesuu", "Reola", "Riisa", "Ristna",
    "Rohuneeme", "Roomassaare", "Roosisaare", "Roostoja", "Ruhnu", "Sõrve", "Sämi", "Särevere",
    "Tallinn-Harku", "Tartu", "Tartu-Tõravere", "Taheva", "Tahkuse", "Tiirikoja", "Toila-Oru", "Tooma",
    "Tudu", "Tuulemäe", "Türi", "Tõlliste", "Tõrva", "Tõrve", "Uue-Lõve", "Vaindloo", "Valga",
    "Valgu", "Vanaküla", "Varangu", "Vasknarva", "Vihterpalu", "Viljandi", "Vilsandi", "Virtsu",
    "Võru", "Väike-Maarja"
]

# Funktsioon Exceli faili varukoopia tegemiseks
def tee_varukoopia():
    varukoopia_nimi = f"backup_{os.path.basename(failinimi)}"
    try:
        shutil.copy(failinimi, os.path.normpath(os.path.join(faili_asukoht, varukoopia_nimi)))
        print(f"Varukoopia loodud: {os.path.normpath(os.path.join(faili_asukoht, varukoopia_nimi))}")
    except Exception as e:
        print(f"Tekkis viga varukoopia tegemisel: {e}")

# Funktsioon faili salvestuskoha valimiseks
def vali_salvestuskoht():
    global faili_asukoht
    salvestuskoht = filedialog.askdirectory()
    if salvestuskoht:
        faili_asukoht = salvestuskoht
    else:
        faili_asukoht = os.getcwd()
    uuenda_failinimi()
    print(f"Faili asukoht on nüüd: {os.path.normpath(faili_asukoht)}")

# Uuenda failinimi vastavalt salvestuskohale
def uuenda_failinimi():
    global failinimi
    failinimi = os.path.normpath(os.path.join(faili_asukoht, "ilm.xlsx"))
    print(f"Fail salvestatakse asukohta: {failinimi}")

    
andmed = []
url = "https://www.ilmateenistus.ee/ilm/ilmavaatlused/vaatlusandmed/tunniandmed/"

# Geckodriveri tee määramine
geckodriver_path = "C:/Program Files/geckodriver/geckodriver.exe"  # Asenda oma WebDriveri tee

# Algse asukoha määramine
faili_asukoht = os.getcwd()

root = Tk()
root.title("Ilmavaatlus")
root.geometry("500x250")

Label(root, text="Sisesta kuupäev (pp.kk.aaaa):").grid(row=0, column=0, padx=10, pady=5)
kuupäev_sisend = Entry(root, width=15)
kuupäev_sisend.grid(row=0, column=1, padx=10, pady=5)
Button(root, text="Vali kuupäev", command=vali_kuupäev).grid(row=0, column=2, padx=10, pady=5)

Label(root, text="Sisesta hommikune kellaaeg (tt:00):").grid(row=1, column=0, padx=10, pady=5)
hommikune_aeg_sisend = Entry(root, width=15)
hommikune_aeg_sisend.grid(row=1, column=1, padx=10, pady=5)

Label(root, text="Sisesta õhtune kellaaeg (tt:00):").grid(row=2, column=0, padx=10, pady=5)
õhtune_aeg_sisend = Entry(root, width=15)
õhtune_aeg_sisend.grid(row=2, column=1, padx=10, pady=5)

Label(root, text="Vali asukoht:").grid(row=3, column=0, padx=10, pady=5)
asukoht_sisend = ttk.Combobox(root, values=asukohad, state='readonly')
asukoht_sisend.grid(row=3, column=1, padx=10, pady=5)


Button(root, text="Tee varukoopia", command=tee_varukoopia).grid(row=4, column=0, padx=10, pady=5)
Button(root, text="Vali salvestuskoht", command=vali_salvestuskoht).grid(row=4, column=1, padx=10, pady=5)

Button(root, text="Käivita", command=käivita).grid(row=5, column=0, columnspan=2, padx=10, pady=20)

# Uuenda failinimi kohe alguses, et määrata vaikesalvestuskoht
uuenda_failinimi()

root.mainloop()
