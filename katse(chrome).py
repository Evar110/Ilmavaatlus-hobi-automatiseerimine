from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service as ChromeService
from bs4 import BeautifulSoup
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Side, Font
from tkinter import Tk, Label, Entry, Button
import os


# funktsioon andmete kogumiseks
def andmete_kogumine(kellaaeg, kogutav, koht):
    """
    :parameeter kellaaeg: Kellaaeg, mida kasutada päringus (näiteks "7:00").
    :parameeter kogutav: Andmed, mida koguda ('hommik' või 'õhtu').
    :parameeter koht: Asukoht, kus andmeid võtab.
    """
    global driver
    try:
        # Avab veebilehe
        driver.get(url)

        # Kuupäeva sisestamine
        date_input = WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.ID, "ood-datepicker")))  # ootab 20s või kuni saab lisada kuupäeva
        date_input.clear()
        date_input.send_keys(kuupäev)
        date_input.send_keys(Keys.RETURN)
        
        # Kellaaja sisestamine
        time_input = WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.ID, "ood-timepicker")))  # ootab 20s või kuni saab lisada kellaaja
        time_input.clear()
        time_input.send_keys(kellaaeg)
        time_input.send_keys(Keys.RETURN)

        # Ootab 20s või kuni leht lõpetab andmete laadimise
        WebDriverWait(driver, 20).until(lambda d: driver.execute_script("return jQuery.active == 0"))

        # Parsib HTML-i BeautifulSoup abil
        html = driver.page_source
        soup = BeautifulSoup(html, "html.parser")
        tabel = soup.find("table", class_="table")

        if tabel:
            read = tabel.find_all("tr")[1:]  # Jätab vahele päise rea
            for rida in read:
                andmete_rida = rida.find_all("td")
                if andmete_rida:
                    asukoht = andmete_rida[0].get_text(strip=True)
                    temperatuur = andmete_rida[1].get_text(strip=True)
                    õhurõhk = andmete_rida[3].get_text(strip=True)
                    tuule_suund = andmete_rida[5].get_text(strip=True)

                    if asukoht == koht:  # Filtreerib valitud asukoha andmed
                        if kogutav == "hommik":
                            print(f"Asukoht: {asukoht}")
                        print(f"Temperatuur: {temperatuur if temperatuur else '-'}")
                        andmed.append(temperatuur if temperatuur else '-')

                        if kogutav == "õhtu":
                            print(f"Õhurõhk: {õhurõhk if õhurõhk else '-'}")
                            print(f"Tuule suund: {tuule_suund if tuule_suund else '-'}")
                            andmed.extend([
                                õhurõhk if õhurõhk else '-',
                                tuule_suund if tuule_suund else '-'
                            ])
                            print("-" * 40)
    except Exception as e:
        print(f"Tekkis viga: {e}")

# funktsioon nädalapäeva saamiseks
def nädalapäev(kuupäev): 
    nädalapäevad = ['E', 'T', 'K', 'N', 'R', 'L', 'P']
    
    kuupäeva_obj = datetime.strptime(kuupäev, "%d.%m.%Y")
    nädalapäeva_indeks = kuupäeva_obj.weekday()
    return f"{kuupäeva_obj.day}.{nädalapäevad[nädalapäeva_indeks]}."

# funktsioon, mis teisendab hektopaskalid mmHg-ks
def hPa_mmHg(hPa):
    mmHg = str(round(float(hPa) * 0.750062)) + "mmHg"
    return mmHg
    

# funktsioon, mis muudab tuule suuna kraad ilmakaareks
def tuule_suund(kraadid):
    ilmakaared = ["Põhja", "Kirde", "Ida", "Kagu", "Lõuna", "Edela", "Lääne", "Loode", "Põhja"]
    suuna_indeks = int((int(kraadid) + 22.5) % 360 // 45)
    return ilmakaared[suuna_indeks]

# Funktsioon, mis kontrollib kas on uus aasta
def uus_aasta(kuupäev):
    if int(kuupäev[:2]) == 1 and int(kuupäev[3:5]) == 1:
        return True
    else:
        return False
    
# funktsioon, mis kontrollib kas on uus kuu
def uus_kuu(kuupäev):
    if int(kuupäev[:2]) == 1:
        return True
    else:
        return False

# funktsioon, mis tagastab kuu ja aasta, mida kasutaja sisestas    
def kuu_ja_aasta(kuupäev):
    järjend = []
    kuud = ['Jaanuar', 'Veebruar', 'Märts', 'Aprill', 'Mai', 'Juuni', 'Juuli', 'August', 'September', 'Oktoober', 'November', 'Detsember']
    kuu = kuud[int(kuupäev[3:5]) - 1]
    järjend.append(kuu)
    aasta = int(kuupäev[6:])
    järjend.append(aasta)
    return järjend
    
# funktsioon, mis kirjutab andmed Exceli faili
def lisa_excelisse(failinimi, andmed, kuupäev):
    """
    Lisab andmed Exceli faili uuele reale, kui faili ei ole, loob selle.
    Lisaks määrab joondamise, stiili ja muud visuaalsed elemendid.
    """
    try:
        # Kuu ja aasta
        kuu_ja_aasta_järjend = kuu_ja_aasta(kuupäev)
        aasta = kuu_ja_aasta_järjend[1]
        kuu = kuu_ja_aasta_järjend[0]

        # Stiilid
        keskele_joondus = Alignment(horizontal="center")
        alläär = Border(bottom=Side(style="thin"))
        suur_font = Font(size=18)

        # Spetsiaalsed read
        päis_rida = ["Kuupäev", "Hommiku t°", "Õhtu t°", "Õhurõhk", "Tuule suund"]
        kuu_rida = ["", "", kuu]
        aasta_rida = ["", "", aasta]

        #Juhul, kui ei ole sellist faili, loob selle
        if not os.path.exists(failinimi):
            fail = Workbook()
            excel = fail.active

            # Lisab aasta
            excel.append(aasta_rida)
            for lahter in excel[excel.max_row]:
                lahter.alignment = keskele_joondus
                lahter.font = suur_font

            # Lisab kuu
            excel.append(kuu_rida)
            for lahter in excel[excel.max_row]:
                lahter.alignment = keskele_joondus

            # Lisab päise
            excel.append(päis_rida)
            for lahter in excel[excel.max_row]:
                lahter.alignment = keskele_joondus
                lahter.border = alläär

            # Lisab andmed
            excel.append(andmed)
            for lahter in excel[excel.max_row]:
                lahter.alignment = keskele_joondus

            # Salvestab andmed
            fail.save(failinimi)

        else:
            fail = load_workbook(failinimi)
            excel = fail.active
            # Kui on uus aasta, lisab uue aasta, kuu ja päise
            if uus_aasta(kuupäev):
                excel.append(aasta_rida)
                for lahter in excel[excel.max_row]:
                    lahter.alignment = keskele_joondus
                    lahter.font = suur_font

                excel.append(kuu_rida)
                for lahter in excel[excel.max_row]:
                    lahter.alignment = keskele_joondus

                excel.append(päis_rida)
                for lahter in excel[excel.max_row]:
                    lahter.alignment = keskele_joondus
                    lahter.border = alläär

            # Kui on uus kuu, lisab uue kuu
            elif uus_kuu(kuupäev):
                excel.append(kuu_rida)
                for lahter in excel[excel.max_row]:
                    lahter.alignment = keskele_joondus

            # Lisab andmed        
            excel.append(andmed)
            for lahter in excel[excel.max_row]:
                lahter.alignment = keskele_joondus
        
            # Salvestab andmed
            fail.save(failinimi)

    # Juhul, kui kasutaja jätab exceli faili lahti, kui programmi käivitab
    except PermissionError:
        print("Viga, ei saanud salvestada faili, kontrolli ega sul exceli fail avatud ei ole")

# Funktsioon, mis kustutab varasemad kogutud andmed, et programm saaks uuesti koguda andmeid, ilma restarti tegematta.
def restart():
    global andmed
    andmed = []

def käivita():
    global kuupäev, hommikune_aeg, õhtune_aeg, asukoht, andmed, driver

    kuupäev = kuupäev_sisend.get()
    hommikune_aeg = hommikune_aeg_sisend.get()
    õhtune_aeg = õhtune_aeg_sisend.get()
    asukoht = asukoht_sisend.get()

    # Kontroll
    if len(kuupäev) != 10 or not hommikune_aeg or not õhtune_aeg or not asukoht:
        print("Palun täida kõik väljad!")
        return
    
    restart()

    # Käivitab Chrome WebDriver
    service = ChromeService(executable_path=geckodriver_path)
    driver = webdriver.Chrome(service=service)

    try:
        # Andmete kogumine hommikul (ainult temperatuur)
        andmete_kogumine(hommikune_aeg, "hommik", asukoht)

        # Andmete kogumine õhtul (temperatuur, õhurõhk, tuule suund)
        andmete_kogumine(õhtune_aeg, "õhtu", asukoht)
    finally:
        driver.quit()

    # Muudab andmed järjendi elemente, et vastaksid tahetud formaadile
    päev_nädalapäev = nädalapäev(kuupäev)
    andmed.insert(0, päev_nädalapäev)
    õhurõhk = andmed[3]
    õhurõhk = õhurõhk.replace(',','.')
    õhurõhk = hPa_mmHg(õhurõhk)
    andmed[3] = õhurõhk
    tuule_kraadid = andmed[4]
    ilmakaar = tuule_suund(tuule_kraadid)
    andmed[4] = ilmakaar

    print("Kogutud andmed:", andmed)

    failinimi = "ilm.xlsx"
    lisa_excelisse(failinimi, andmed, kuupäev)

    print(f"Andmed salvestatud faili {failinimi}.")
    
andmed = []
url = "https://www.ilmateenistus.ee/ilm/ilmavaatlused/vaatlusandmed/tunniandmed/"

# Geckodriveri tee määramine
geckodriver_path = "C:/Program Files/geckodriver/geckodriver.exe"  # Asenda oma WebDriveri tee

root = Tk()
root.title("Ilmavaatlus")
root.geometry("400x300")

Label(root, text="Sisesta kuupäev (pp.kk.aaaa):").pack(pady=5)
kuupäev_sisend = Entry(root, width=30)
kuupäev_sisend.pack(pady=5)

Label(root, text="Sisesta hommikune kellaaeg (tt:00):").pack(pady=5)
hommikune_aeg_sisend = Entry(root, width=30)
hommikune_aeg_sisend.pack(pady=5)

Label(root, text="Sisesta õhtune kellaaeg (tt:00):").pack(pady=5)
õhtune_aeg_sisend = Entry(root, width=30)
õhtune_aeg_sisend.pack(pady=5)

Label(root, text="Sisesta asukoht:").pack(pady=5)
asukoht_sisend = Entry(root, width=30)
asukoht_sisend.pack(pady=5)

# Nupp protsessi käivitamiseks
Button(root, text="Käivita", command=käivita).pack(pady=20)

root.mainloop()